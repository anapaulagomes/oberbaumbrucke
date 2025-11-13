import marimo

__generated_with = "0.16.4"
app = marimo.App(width="medium")


@app.cell
def _():
    import marimo as mo
    import polars as pl
    from openpyxl import load_workbook
    return load_workbook, pl


@app.cell
def _(load_workbook):
    workbook = load_workbook('data/mappings/CID10-v2019-lista-codigos.xlsx')
    return (workbook,)


@app.cell
def _(workbook):
    workbook.sheetnames
    return


@app.cell
def _(workbook):
    def get_codes(sheet_index, col, row):
        for _row in workbook.worksheets[sheet_index].iter_rows(max_col=col, min_row=row):
            if _row[0].value:
                yield _row[0].value.replace(".", "").strip(), _row[1].value.strip()
    return (get_codes,)


@app.cell
def _(get_codes):
    removed = list(get_codes(3, 2, 3))
    return (removed,)


@app.cell
def _(pl, removed):
    df = pl.DataFrame(removed, schema=["code", "description"])
    df
    return (df,)


@app.cell
def _(df):
    df.write_csv("data/mappings/cid-10-bra-removed.csv")
    return


if __name__ == "__main__":
    app.run()
